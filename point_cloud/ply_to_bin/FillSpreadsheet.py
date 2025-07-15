#!/usr/bin/env python3
#--------------------------------------------------------------------------------
# Copyright (c) 2025 InterDigital CE Patent Holdings
# 
# Licensed under the License terms and conditions for use, reproduction, and
# distribution of 5G-MAG software (the “License”).  You may not use this file
# except in compliance with the License.  You may obtain a copy of the License at
# https://www.5g-mag.com/reference-tools.  Unless required by applicable law or
# agreed to in writing, software distributed under the License is distributed on
# an “AS IS” BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
# or implied.
# 
# See the License for the specific language governing permissions and limitations
# under the License.
#--------------------------------------------------------------------------------

import os
import sys
import argparse
import zipfile
import openpyxl

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def getCurrentDir():
  return os.path.dirname( os.path.realpath(__file__) )

# Set configurations
seqOrder = ["mitch", "juggleSoccer", "nathalie", "henry", "aliyah"]
numRate =  5
startLine = 5 #startLineInEmptyXls
strCond = [{'name': 'C2 lossy RA', 'dataColumn': [15, 31], 'nbSourcePoints': 4, 'frameColumn': 6}]

lossyColumn = ['NbOutputPoints', 'MeanOutputPoints', 'MeanDuplicatePoints', 'TotalBitstreamBits', 'geometryBits', 'metadataBits', 'attributeBits', 'D1Mean', 'D2Mean', 'LumaMean', 'CbMean', 'CrMean', 'PCQM', 'SelfEncoderRuntime','ChildEncoderRuntime', 'SelfDecoderRuntime','ChildDecoderRuntime']

def parseArgs():
  global parser
  source = getCurrentDir() + "/empty.xlsm"
  parser = argparse.ArgumentParser(description="Fill .xlsm file")
  # clang-format off
  parser.add_argument("--source", help="input template spreadsheet path (xlsm)", default=source,       type=str)
  parser.add_argument("--save",   help="created spreadsheet path (xlsm)",        default="test.xlsm",  type=str)
  parser.add_argument("--tested", help="input CSV path (csv)",                   default="test.csv",   type=str)
  parser.add_argument("--frame",  help="Number of frames",                       default=-1,           type=int)
  # clang-format one
  return parser.parse_args()

# DataFrame to read our input CS file
def readCsv(filename):
  from csv import DictReader
  with open( filename, 'r') as read_obj:
    data = []
    for row in DictReader(read_obj):
      data.append( row )
  return data

# Print sequences
def printData(data):
  #print("data:", data)
  print("")
  for line in data:
    print(" - Cond = %s seq = %s rate = %s: %s " %
      (line["CondId"], line["SeqId"], line["RateId"], line))
    print("")
  print("")

# Get results of one test
def get(array, condId, seqId, rateId):
  #print("get - condId=", condId, "seqId=", seqId, "rateId=", rateId)
  for el in array:
    if (el['SeqId'] == seqId and el['RateId'] == rateId):
      return (el)
  return ()

#  Update the results of one experiment
def update(wb, path, nbSseqToHandle, nbRateToHandle, testId=0):
    data = readCsv(path)
    #printData(data)
    condId=0
    ws = wb[strCond[condId]['name']]
    columnId = strCond[condId]['dataColumn'][testId]
    for idx, seqId in enumerate(seqOrder):
        if (idx > nbSseqToHandle-1):
            #print("reach max seq to handle in worksheet : ", seqId)
            break;
        else:
            for rate in range(0, numRate):
                
                if (rate > nbRateToHandle-1):
                    #print("reach max rate to handle in worksheet : ", rate)
                    break;

                minSeqId=1
                el = get(data, condId, "".join(["S",str(idx+minSeqId)]), "".join(["R%02d" % (rate+1)]))

                #update nbSourcePoints
                ws.cell( startLine + idx * startLine + rate, strCond[condId]['nbSourcePoints']).value = int(el['NbInputPoints'])
                #update nbFrame
                ws.cell( startLine + idx * startLine + rate, strCond[condId]['frameColumn']).value = int(el['nbFrame'])

                #update data
                for i in range(0, len(lossyColumn)):
                  cell = ws.cell(startLine + idx * startLine + rate, columnId + i)
                  value = float(el[lossyColumn[i]])
                  cell.value = '' if value<0 else value if i>1 else int(value)    
              
# Fill xlsm file
def fillXlsm(source, output, tested, frame, nbSseqToHandle, nbRateToHandle):
  wb = load_workbook(source, read_only=False, keep_vba=True)
  update(wb, tested, nbSseqToHandle, nbRateToHandle, 0)
  wb.save(output)
  
##################
def main():

    try:
        # Parse arguments
        args = parseArgs()
        frameworkUtils.printArgs(args)  

        # Check options
        if args.frame <= 0:
            raise ValueError(args.frame, "Error: number of frame must be set")
        if not os.path.isfile( args.source ) and not os.path.exists( args.source ):
            raise ValueError(args.source, "Error: source xlsm file not exist: %s")

        # Fill xlsm
        fillXlsm( args.source, args.save, args.tested, args.frame )
        print("%s create %s " % ( os.path.basename(__file__), args.save) )

    except Exception as e:
        print (frameworkUtils.RED + "Exception:", e, frameworkUtils.ENDC, flush=True)
        parser.print_help(sys.stderr)
        sys.exit(e.returncode)
        
##################
if __name__ == "__main__":
    
    main()
    
